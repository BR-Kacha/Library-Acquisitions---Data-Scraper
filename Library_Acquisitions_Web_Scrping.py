from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait 
import openpyxl
from selenium.webdriver.common.by import By
from pynput.keyboard import Key, Controller
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import csv
import pandas as pd
import xlsxwriter
import time
import datetime

time_obj_now = datetime.datetime.now()
print ("Start date and time : ")
print (time_obj_now.strftime("%Y-%m-%d %H:%M:%S"))

wb = openpyxl.load_workbook("Book Details.xlsx")
sheet = wb.active

unique_isbn=[]
lib_data_dict = {"ISBN":["ISBN 13","ISBN 10","Title","Author","Publisher"]}

keyboard = Controller()
driver = webdriver.Chrome("D:\chromedriver-win64\chromedriver-win64\chromedriver.exe")  #Update your Chrome Driver path here
driver.get("https://www.goodreads.com/ap/signin?openid.return_to=https%3A%2F%2Fwww.goodreads.com%2Fap-handler%2Fregister&prevRID=JHAW05ZHTS3FPZ1T0EMG&openid.identity=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0%2Fidentifier_select&openid.assoc_handle=amzn_goodreads_web_na&openid.mode=checkid_setup&siteState=ebe6b7f4ea22ec4e80e133fe98da8bc2&language=en_US&openid.claimed_id=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0%2Fidentifier_select&pageId=amzn_goodreads_web_na&openid.ns=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0")

mail_txtbox = driver.find_element(By.XPATH, '//*[@id="ap_email"]')
mail_txtbox.send_keys("") #Enter your Gooreads.com account e-mail between ""

password_txtbox = driver.find_element(By.XPATH, '//*[@id="ap_password"]')
password_txtbox.send_keys("") #Enter your Gooreads.com account password between ""

signup_button = driver.find_element(By.XPATH,'//*[@id="signInSubmit"]').click()

for i in range(2,sheet.max_row+1):
    cellbox = sheet.cell(row = i, column = 1)
    isbn = cellbox.value
    
    if isbn == None:
        continue
   #time.sleep(2)
    
    elif isbn in unique_isbn:
        data = lib_data_dict[isbn]
        #write data to excel file
        df1 = pd.DataFrame(data)
        df1 = df1.transpose()
        with pd.ExcelWriter("Book Details 14-09-2023.xlsx", engine="openpyxl",mode = "a", if_sheet_exists="overlay") as writer:
            df1.to_excel(writer, startrow = i-1, startcol = 2, index = False, header = False)
        
    else:
        #try:
            if i==2:
                searchbar_1 = driver.find_element(By.XPATH,'//*[@id="bodycontainer"]/div/div[2]/div/header/div[2]/div/div[2]/form/input[1]')
                searchbar_1.send_keys(str(isbn))
                search_button = driver.find_element(By.XPATH,'//*[@id="bodycontainer"]/div/div[2]/div/header/div[2]/div/div[2]/form/button').click()
            else:
                try:
                    searchbar = driver.find_element(By.XPATH,'//*[@id="Header"]/div[2]/div[2]/section/form/input[1]')
                    search_button = driver.find_element(By.XPATH,'//*[@id="Header"]/div[2]/div[2]/section/form/button')
                except:
                    searchbar = driver.find_element(By.CSS_SELECTOR,'#bodycontainer > div.siteHeader > div > header > div.siteHeader__topLine.gr-box.gr-box--withShadow > div > div.searchBox.searchBox--navbar > form > input.searchBox__input.searchBox__input--navbar')
                    search_button = driver.find_element(By.CSS_SELECTOR,'#bodycontainer > div.siteHeader > div > header > div.siteHeader__topLine.gr-box.gr-box--withShadow > div > div.searchBox.searchBox--navbar > form > button')
                
                #entering the book name in searchbar
                searchbar.send_keys(str(isbn))
                #clicking the search button
                search_button.click()
                try:
                    #clicking the found book
                    book = driver.find_element(By.XPATH,'//*[@id="bodycontainer"]/div[3]/div[1]/div[2]/div[2]/table/tbody/tr/td[2]/a').click()
                except:
                    #if whole book is not available on website
                    data_lst = ["NOT FOUND","NOT FOUND","NOT FOUND","NOT FOUND","NOT FOUND"]
                    df2 = pd.DataFrame(data_lst)
                    df2 = df2.transpose()
                    with pd.ExcelWriter("Book Details.xlsx", engine="openpyxl",mode = "a", if_sheet_exists="overlay") as writer:
                        df2.to_excel(writer, startrow = i-1,startcol = 2 ,index = False, header = False)
                    unique_isbn.append(isbn)
                    lib_data_dict[isbn] = data_lst
                
            #WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '//*[@id="__next"]/div[2]/main/div[1]/div[2]/div[1]/div[2]/div[6]/div/div'))).click()
            try:
                more_detail_button = driver.find_element(By.CSS_SELECTOR,'#__next > div.PageFrame.PageFrame--siteHeaderBanner > main > div.BookPage__gridContainer > div.BookPage__rightColumn > div.BookPage__mainContent > div.BookPageMetadataSection > div.BookDetails > div > div > button').click()
                # more_detail_button = driver.find_element(By.XPATH,'//*[@id="__next"]/div[2]/main/div[1]/div[2]/div[1]/div[2]/div[6]/div/div').click()
            except:
                pass
            # img_click = driver.find_element(By.XPATH,'//*[@id="__next"]/div[2]/main/div[1]/div[1]/div/div[1]/div/div/div/div/div/div/img').click()
            try:
                try:
                    book_name = driver.find_element(By.CSS_SELECTOR,'#__next > div.PageFrame.PageFrame--siteHeaderBanner > main > div.BookPage__gridContainer > div.BookPage__rightColumn > div.BookPage__mainContent > div.BookPageTitleSection > div.BookPageTitleSection__title > h1').text
                    # book_name = driver.find_element(By.XPATH,'//*[@id="__next"]/div[2]/main/div[1]/div[2]/div[1]/div[1]/div[1]/h1').text
                except:
                    book_name = "NOT FOUND"   
                try:
                    author = driver.find_element(By.CSS_SELECTOR,'#__next > div.PageFrame.PageFrame--siteHeaderBanner > main > div.BookPage__gridContainer > div.BookPage__rightColumn > div.BookPage__mainContent > div.BookPageMetadataSection > div.BookPageMetadataSection__contributor > h3 > div > span:nth-child(1) > a > span').text
                # author = driver.find_element(By.XPATH,'//*[@id="__next"]/div[2]/main/div[1]/div[2]/div[1]/div[2]/div[1]/h3/div/span[1]/a/span').text
                except:
                    author = "NOT FOUND"
                try:
                    isbn_13 = driver.find_element(By.CSS_SELECTOR,'#__next > div.PageFrame.PageFrame--siteHeaderBanner > main > div.BookPage__gridContainer > div.BookPage__rightColumn > div.BookPage__mainContent > div.BookPageMetadataSection > div.BookDetails > div > span:nth-child(2) > div.BookDetails__list > span > div > dl > div:nth-child(3) > dd > div > div.TruncatedContent__text.TruncatedContent__text--small').text
                    isbn_10 = isbn_13[23:33]
                    isbn_13 = isbn_13[0:13]
                except:
                    isbn_13 = "NOT FOUND"
                    isbn_10 = "NOT FOUND"
                try:    
                    publish = driver.find_element(By.CSS_SELECTOR,'#__next > div.PageFrame.PageFrame--siteHeaderBanner > main > div.BookPage__gridContainer > div > div.BookPage__mainContent > div.BookPageMetadataSection > div.BookDetails > div > span:nth-child(2) > div.BookDetails__list > span > div > dl > div:nth-child(2) > dd > div > div.TruncatedContent__text.TruncatedContent__text--small').text
                except:
                    publish = "NOT FOUND"
            except:
                continue
            data_lst = [isbn_13,isbn_10,book_name,author,publish]
            
            df2 = pd.DataFrame(data_lst)
            df2 = df2.transpose()
    
            with pd.ExcelWriter("Book Details.xlsx", engine="openpyxl",mode = "a", if_sheet_exists="overlay") as writer:
                df2.to_excel(writer, startrow = i-1,startcol = 2 ,index = False, header = False)
            
            unique_isbn.append(isbn)
            lib_data_dict[isbn] = data_lst
       # except:
          #  pass
time_obj_now = datetime.datetime.now()

print ("Completion date and time : ")
print (time_obj_now.strftime("%Y-%m-%d %H:%M:%S"))
        
driver.close()


